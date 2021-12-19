import sqlite3
import csv
import shutil
from openpyxl import Workbook, load_workbook
import pandas as pd
import numpy as np
import datetime
import shutil
import json

# effort - package,date,engineerName,startTime,endTime,workingHours,overtime,location,worklog
# travel - engineerName,date,type,city,description,invoiceType,price
def excel_to_json(workbook_name, sheet_name, header_row, row_start, row_end, col_start, col_end, base_month):
    try:
        # metadata dict
        t_meta = {'workbook_name': workbook_name, 'sheet_name': sheet_name, 'header_row': header_row, 'row_start': row_start, 'row_end':row_end, 'col_start':col_start, 'col_end': col_end, 'base_month': base_month}
        t_wb = load_workbook(workbook_name, read_only=True, data_only=True)
        t_ws = t_wb[sheet_name]
        current_row = 0
        table_header = []
        data_dict = {}
        # get table header dict
        for col in range(col_start, col_end + 1):
            table_header.append(t_ws.cell(row=header_row, column=col).value)

        # get table header dict
        for row in t_ws.iter_rows(min_row=row_start, max_row=row_end, min_col=col_start, max_col=col_end, values_only=True):
            # print(type(row), len(row), row)
            current_row += 1
            data_dict[current_row] = row

        all_dict = {'metadata':t_meta, 'header': table_header, 'data': data_dict}
        json_file_name = 'all_dict.json'
        with open(json_file_name, 'w') as json_file:
            json.dump(all_dict, json_file, ensure_ascii=False, indent=4)
            
        return json_file_name
    except Exception as e:
        return '<h1>error when reading excel</h1><p>{}</p>'.format(e)

def check_data(file_to_check, nt_name):
    import re
    with open(file_to_check, 'r') as j:
        json_data = json.load(j)

    # when nt_names == '_all_'
    f_nt_check = True
    if nt_name == '_all_':
        f_nt_check = False

    # checking blank
    rows_of_blank_record = []
    rows_of_overdue_record = []
    rows_of_error_record = []
    base_month = json_data['metadata']['base_month'][:7].replace('-', '')
    # print(base_month)
    col_num_of_nt_name = json_data['header'].index('Responsible Sales')
    for key, value in json_data['data'].items():
        if f_nt_check == False or value[col_num_of_nt_name] == nt_name:
            if None in value[-4:]:
                rows_of_blank_record.append(key)
            try:
                if int(re.sub(r'[-.]', '', value[-2])) < int(base_month):
                    rows_of_overdue_record.append(key)
            except:
                rows_of_error_record.append(key)

    dict_to_return = {'rows_of_blank_record':rows_of_blank_record, 'rows_of_overdue_record':rows_of_overdue_record, 'rows_of_error_record': rows_of_error_record}
    return dict_to_return

def json_to_python(file_to_load):
    from werkzeug.utils import secure_filename
    with open(file_to_load, 'r') as j:
        json_data = json.load(j)

    # # return json_data
    # print(type(json_data['header']))
    # print(json_data['header'].index('Responsible Sales'))
    # print(type(json_data['data']))
    # distinct_nt_id = {'#'}
    # col_num_of_nt_id = json_data['header'].index('Responsible Sales')
    # for key, value in json_data['data'].items():
    #     print(type(key))
    #     distinct_nt_id.add(value[col_num_of_nt_id])
    
    # print(type(distinct_nt_id))
    nt_id = 'CUI Jimmy (ED/SCN-C)'
    distinct_nt_id = {'All':'All'}
    col_num_of_nt_name = json_data['header'].index('Responsible Sales')

    for key, value in json_data['data'].items():
        distinct_nt_id[secure_filename(value[col_num_of_nt_name])] = value[col_num_of_nt_name]
    
    print(distinct_nt_id)
    # data_for_jinja = json_data['data']
    # for key, value in json_data['data'].items():
    #     distinct_nt_id.add(value[col_num_of_nt_id])
    # if nt_id == '#':
    #     data_for_jinja = json_data['data']
    # else:
    #     data_for_jinja = dict()
    #     for key, value in json_data['data'].items():
    #         if value[col_num_of_nt_id] == nt_id:
    #             data_for_jinja[key] = value
    # print(data_for_jinja)
    # for key, value in json_data['data'].items():
    #     print(key, value)


def dictionary(property, key, direction):
    dict = [
        ['package', 'CADA/CAPA', '1'],
        ['package', 'PPE', '2'],
        ['invoiceType', 'General', '1'],
        ['invoiceType', 'e-Invoice', '2'],
        ['invoiceType', 'VAT', '3'],
        ['type', 'Accomodation', '1'],
        ['type', 'Meals', '2'],
        ['type', 'Transp.-Flight/Train', '3'],
        ['type', 'Transp.-Taxi/Toll/Bus/Metro', '4']
    ]
    value2return = -1
    if (direction == 1):
        for item in dict:
            if item[0] == property and item[1] == key: value2return = item[2]
    else:
        for item in dict:
            if item[0] == property and item[2] == key: value2return = item[1]
    return value2return

def csv2database(filename, tablename):
    conn = sqlite3.connect("audi.sqlite")
    cur = conn.cursor()
    with open(filename, newline='') as csvfile:
        spamreader = csv.reader(csvfile,delimiter=',')
        line = 0
        for row in spamreader:
            if(line == 0): 
                line += 1
                continue
            if(tablename == 'effort'):
                int_package = dictionary('package', row[0], 1)
                cur.execute('INSERT INTO effort (package,date,engineerName,startTime,endTime,workingHours,overtime,location,worklog) VALUES (?,?,?,?,?,?,?,?,?)', (int_package,row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8]))
            if(tablename == 'travel'):
                int_type = dictionary('type', row[2], 1)
                int_invoiceType = dictionary('invoiceType', row[5], 1)
                cur.execute('INSERT INTO travel (engineerName,date,type,city,description,invoiceType,price) VALUES (?,?,?,?,?,?,?)', (row[0],row[1],int_type,row[3],row[4],int_invoiceType,row[6]))
        conn.commit()
        conn.close()

def records2excel(table, engineerName, month, interested_columns):
    conn = sqlite3.connect('audi.sqlite')
    cur = conn.cursor()
    query_sql = 'SELECT * FROM ' + table + ' WHERE engineerName=? AND strftime("%m", date)=?'
    data = cur.execute(query_sql, (engineerName, month)).fetchall()
    table_header = [item[0] for item in cur.description]
    conn.close()
    # convert number to string
    data_str = int2str(data, table)
    # DataFrame it
    df = pd.DataFrame(data_str, columns=table_header)
    # another DataFrame drop out uninteresting column
    uninteresting_columns = [item for item in table_header if item not in interested_columns]
    df = df.drop(columns=uninteresting_columns)
    # convert string to float: effort/workingHours,overtime;travel/price
    # print(table_header)
    # print(df.head())
    if(table=='effort'):
        df['overtime'] = df['overtime'].astype('float')
        df['workingHours'] = df['workingHours'].astype('float')
    elif(table=='travel'):
        df['price'] = df['price'].astype('float')
    # add Total row
    df.loc['total'] = df.select_dtypes(np.number).sum()
    # excel_id = random.randrange(99,1000)
    try:
        engineerName = engineerName.replace(' ', '')
        tmp_str = '-'
        engineerName = tmp_str.join(engineerName.split('/'))
    except:
        print('engineerName normal')
    # excel_name = str(excel_id) + '_' + table + '_report_' + engineerName + month + '.xlsx'
    excel_name = table + '_report_' + engineerName + '_' + month + '.xlsx'
    df.to_excel('persist/excels/' + excel_name)
    # write title and No into excel
    wb = load_workbook('persist/excels/' + excel_name)
    ws = wb.active
    ws.insert_rows(1, 2)
    ws.cell(row=3,column=1,value='NO')
    title = table + ' report for ' + engineerName + ' of month ' + month
    ws.cell(row=1,column=1,value=title)
    wb.save('persist/excels/' + excel_name)
    return excel_name

def excel_summary(tableName, month):
    conn = sqlite3.connect('audi.sqlite')
    cur = conn.cursor()
    if tableName == 'effort':
        columns_effort_interested = ['package','date','startTime','endTime','workingHours','overtime','location','worklog']
        query_sql = 'SELECT DISTINCT engineerName FROM ' + tableName + ' WHERE strftime("%m", date)=?'
        distinct_engineerNames = cur.execute(query_sql, (month,)).fetchall()
        for engineer in distinct_engineerNames:
            # engineer = str(engineer)
            engineer = engineer[0]
            query_sql = 'SELECT * FROM ' + tableName + ' WHERE engineerName=? AND strftime("%m", date)=?'
            data = cur.execute(query_sql, (engineer, month)).fetchall()
            table_header = [item[0] for item in cur.description]
            data_str = int2str(data, tableName)
            df = pd.DataFrame(data_str, columns=table_header)
            # another DataFrame drop out uninteresting column
            uninteresting_columns = [item for item in table_header if item not in columns_effort_interested]
            df = df.drop(columns=uninteresting_columns)
            df['overtime'] = df['overtime'].astype('float')
            df['workingHours'] = df['workingHours'].astype('float')
            df.loc['total'] = df.select_dtypes(np.number).sum()
            try:
                engineer = engineer.replace(' ', '')
                tmp_str = '-'
                engineer = tmp_str.join(engineer.split('/'))
            except:
                print('engineerName normal')
            with pd.ExcelWriter('persist/effortSummaryTemplate.xlsx', mode='a') as writer:
                df.to_excel(writer, sheet_name=engineer)
    elif tableName == 'travel':
        columns_travel_interested = ['date','type','city','description','invoiceType','price']
        query_sql = 'SELECT DISTINCT engineerName FROM ' + tableName + ' WHERE strftime("%m", date)=?'
        distinct_engineerNames = cur.execute(query_sql, (month,)).fetchall()
        for engineer in distinct_engineerNames:
            # engineer = str(engineer)
            engineer = engineer[0]
            query_sql = 'SELECT * FROM ' + tableName + ' WHERE engineerName=? AND strftime("%m", date)=?'
            data = cur.execute(query_sql, (engineer, month)).fetchall()
            table_header = [item[0] for item in cur.description]
            data_str = int2str(data, tableName)
            df = pd.DataFrame(data_str, columns=table_header)
            # another DataFrame drop out uninteresting column
            uninteresting_columns = [item for item in table_header if item not in columns_travel_interested]
            df = df.drop(columns=uninteresting_columns)
            df['price'] = df['price'].astype('float')
            df.loc['total'] = df.select_dtypes(np.number).sum()
            try:
                engineer = engineer.replace(' ', '')
                tmp_str = '-'
                engineer = tmp_str.join(engineer.split('/'))
            except:
                print('engineerName normal')
            with pd.ExcelWriter('persist/travelSummaryTemplate.xlsx', mode='a') as writer:
                df.to_excel(writer, sheet_name=engineer)

def pivot_travel(month, return_type):
    conn = sqlite3.connect('audi.sqlite')
    cur = conn.cursor()
    query_sql = 'SELECT * FROM travel WHERE strftime("%m", date)=?'
    data = cur.execute(query_sql, (month, )).fetchall()
    table_header = [item[0] for item in cur.description]
    conn.close()
    # convert number in records to string
    data_str = int2str(data, 'travel')
    # header + data -> DataFrame
    try:
        df = pd.DataFrame(data_str, columns=table_header)
        df['price'] = df['price'].astype('float')
        pivot_table = pd.pivot_table(df, values='price', index='engineerName', columns='type', aggfunc=np.sum, fill_value=0, margins=True, margins_name='Total')
        if(return_type == 'excel'):
            filename = 'persist/pivot_travel_' + month + '.xlsx'
            pivot_table.to_excel(filename)
            return filename
        else:
            filename = 'pivot_travel_' + month + '.html'
            pivot_table.to_html('templates/' + filename)
            return filename
    except:
        print('')

def int2str(collection, tableName):
    tmp_collection = np.asarray(collection)
    if(tableName == 'effort'):
        for item in tmp_collection:
            item[1] = dictionary('package', item[1], -1)
    elif(tableName == 'travel'):
        for item in tmp_collection:
            item[3] = dictionary('type', item[3], -1)
            item[6] = dictionary('invoiceType', item[6], -1)

    return tmp_collection

def excel_effort_summary(month):
    # make a copy of template
    now = str(datetime.datetime.now())[:19]
    now = now.replace(":","_")
    now = now.replace(" ", "_")
    template_path="excel/effortSummaryTemplate.xlsx"
    filename="excel/effortSummary" + month + '_' + str(now) + ".xlsx"
    shutil.copy(template_path,filename)

    columns_effort_interested = ['engineerName', 'date', 'startTime', 'endTime', 'workingHours', 'worklog']
    conn = sqlite3.connect('audi.sqlite')
    cur = conn.cursor()
    query_sql = 'SELECT DISTINCT engineerName FROM effort WHERE strftime("%m", date)=?'
    distinct_engineerNames = cur.execute(query_sql, (month,)).fetchall()
    wb = load_workbook(filename)
    for engineer in distinct_engineerNames:
        engineer = engineer[0]
        query_sql = 'SELECT * FROM effort WHERE engineerName=? AND strftime("%m", date)=?'
        data = cur.execute(query_sql, (engineer, month)).fetchall()
        table_header = [item[0] for item in cur.description]
        data_str = int2str(data, 'effort')
        df = pd.DataFrame(data_str, columns=table_header)
        # another DataFrame drop out uninteresting column
        uninteresting_columns = [item for item in table_header if item not in columns_effort_interested]
        df = df.drop(columns=uninteresting_columns)
        df['workingHours'] = df['workingHours'].astype('float')
        myTotal = df.select_dtypes(np.number).sum()
        myShape = df.shape
        try:
            engineer = engineer.replace(' ', '')
            tmp_str = '-'
            engineer = tmp_str.join(engineer.split('/'))
        except:
            print('engineerName normal')
    
        ws = wb.copy_worksheet(wb['Project Budget'])
        ws.title = engineer
        ws.insert_rows(10, myShape[0])
        current_row = 10
        for index, row in df.iterrows():
            ws.cell(row=current_row, column=1, value=row['engineerName'])
            ws.cell(row=current_row, column=2, value=row['date'])
            ws.cell(row=current_row, column=3, value=row['startTime'])
            ws.cell(row=current_row, column=4, value=row['endTime'])
            ws.cell(row=current_row, column=5, value=row['workingHours'])
            ws.cell(row=current_row, column=6, value=row['worklog'])    
            current_row += 1
        ws.cell(row=current_row,column=5,value=float(myTotal))
        myMonth = '2021-' + month
        ws.cell(row=7, column=2, value=myMonth)
    wb.remove(wb['Project Budget'])
    wb.save(filename)
    return filename

def exceL_travel_summary(month):
    # make a copy of template
    now = str(datetime.datetime.now())[:19]
    now = now.replace(":","_")
    now = now.replace(" ", "_")
    template_path="excel/travelSummaryTemplate.xlsx"
    filename="excel/travelSummary" + month + '_' + str(now) + ".xlsx"
    shutil.copy(template_path,filename)

    conn = sqlite3.connect('audi.sqlite')
    cur = conn.cursor()
    columns_travel_interested = ['date','type','city','description','invoiceType','price']
    query_sql = 'SELECT DISTINCT engineerName FROM travel WHERE strftime("%m", date)=?'
    distinct_engineerNames = cur.execute(query_sql, (month,)).fetchall()
    # append sheet for each engineer
    for engineer in distinct_engineerNames:
        engineer = engineer[0]
        query_sql = 'SELECT * FROM travel WHERE engineerName=? AND strftime("%m", date)=?'
        data = cur.execute(query_sql, (engineer, month)).fetchall()
        table_header = [item[0] for item in cur.description]
        data_str = int2str(data, 'travel')
        df = pd.DataFrame(data_str, columns=table_header)
        # another DataFrame drop out uninteresting column
        uninteresting_columns = [item for item in table_header if item not in columns_travel_interested]
        df = df.drop(columns=uninteresting_columns)
        df['price'] = df['price'].astype('float')
        df.loc['total'] = df.select_dtypes(np.number).sum()
        try:
            engineer = engineer.replace(' ', '')
            tmp_str = '-'
            engineer = tmp_str.join(engineer.split('/'))
        except:
            print('engineerName normal')
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=engineer)
    # write summary sheet
    conn = sqlite3.connect('audi.sqlite')
    cur = conn.cursor()
    myQuery = 'SELECT engineerName, type, price FROM travel WHERE strftime("%m", date)=?'
    result = cur.execute(myQuery, (month,)).fetchall()
    df = pd.DataFrame(result, columns=['engineerName', 'type', 'price'])
    pivot_tabele = pd.pivot_table(df, values='price', index='engineerName', columns='type', aggfunc=np.sum, fill_value=0, margins=True, margins_name='Total')
    myShape = pivot_tabele.shape
    wb = load_workbook(filename)
    ws = wb['Project Budget']
    myMonth = '2021-' + month
    ws.cell(row=6, column=2, value=myMonth)
    ws.insert_rows(9, myShape[0] - 1)
    fake_index = 1
    start_row = 8 # 9 -1
    for index, row in pivot_tabele.iterrows():
        if fake_index == myShape[0]:
            ws.cell(row=start_row + fake_index, column=3,value=float(row[4]))
            ws.cell(row=start_row + fake_index, column=4,value=float(row[3]))
            ws.cell(row=start_row + fake_index, column=5,value=float(row[2]))
            ws.cell(row=start_row + fake_index, column=6,value=float(row[1]))
            ws.cell(row=start_row + fake_index, column=7,value=float(row['Total']))
            break
        ws.cell(row=start_row + fake_index, column=1, value=fake_index)
        ws.cell(row=start_row + fake_index, column=2, value=row.name)
        ws.cell(row=start_row + fake_index, column=3, value=row[4])
        ws.cell(row=start_row + fake_index, column=4, value=row[3])
        ws.cell(row=start_row + fake_index, column=5, value=row[2])
        ws.cell(row=start_row + fake_index, column=6, value=row[1])
        ws.cell(row=start_row + fake_index, column=7, value=row['Total'])
        fake_index += 1
    wb.save(filename)
    return filename

if __name__ == '__main__':
    # readCSV('./0effort.csv', 'effort')
    # readCSV('./0travel.csv', 'travel')
    # readDB()
    # write2excel(1,2)
    # months()
    # print(dictionary('package', 1, -1))
    # data sample
    rows = [[1,'2021-07-05','Transp.-Taxi/Toll/Bus/Metro','Beijing','bullshit','General',219], [2,'2021--07-06','Meals','Beijing','some words','General',114]]
    rows2 = [[1,'PPE','2021-08-11','18:00:00','19:00:00','8.5','5.5', 'bj', 'do abc'], [1,'PPE','2021-08-11','18:00:00','19:00:00','8.5','5.5', 'bj', 'do abc']]
    # print(write2excel('travel', 'whoknows', '07', rows))
    # print(write2excel('effort', 'park', '08', rows2))
    columns_effort_interested = ['package','date','startTime','endTime','workingHours','overtime','location','worklog']
    columns_travel_interested = ['date','type','city','description','invoiceType','price']
    # print(pivot_data2excel('effort', '唐野', '07', columns_effort_interested))
    # print(pivot_data2excel('travel', '唐野', '07', columns_travel_interested))
    # print(pivot_travel('07', 'excel'))
    # print(pivot_travel('07', 'html'))
    # excel_summary('effort', '07')
    # excel_summary('travel', '07')
    # read_some_excel('Sample RBCC_Samos_202108_COEM-20210924.xlsx', 'Detail', 5, 10, 1, 33, 4)
    # json_to_python('all_dict.json')
    # check_data('all_dict.json', 'CUI Jimmy (ED/SCN-C)')
    # check_data('all_dict.json', 'Lv Ping (ED/SCN-C)')
    check_data('all_dict.json', 'Wang Hongmin (ED/SCN-C)')
    check_data('all_dict.json', '_all_')