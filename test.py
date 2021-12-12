from openpyxl import Workbook
from openpyxl import load_workbook

def writeExcel(engineerName, month):
    header_effort = ['package','date','startTime','endTime','workingHours','overtime','location','worklog']
    header_travel = ['date','type','city','description','invoiceType','price']
    data_effort_str = [
        [1, 'CADA/CAPA','2021-07-26','8:55:00','18:10:00','8.0','0.0','Bosch/sgh','do abc'],
        [2, 'CADA/CAPA','2021-07-26','8:55:00','18:10:00','8.0','0.0','Bosch/sgh','do cde']
    ]
    data_travel_str = [
        [1, '2021-07-22','Meals','Beijing','','e-Invoice','85.0'],
        [2, '2021-07-22','Meals','Beijing','','e-Invoice','85.0']
    ]

    wb = load_workbook('template.xlsx')
    ws = wb['effort']
    # printe Title
    title = "Effort report for " + engineerName + ' of month ' + month
    ws.cell(row=3, column=2, value=title)
    # print Header
    for i in range(0,len(header_effort)):
        ws.cell(row=5, column=i+1, value=header_effort[i])

    # print Body
    for x in range(0,len(data_effort_str)):
        ws.insert_rows(6+x)
        for y in range(1,len(header_effort)+1):
            ws.cell(row=x+6, column=y, value=data_effort_str[x][y])

    # print Footer
    total_working = 0
    total_overtime = 0
    for item in data_effort_str:
        total_working += float(item[5])
        total_overtime += float(item[6])
    ws.cell(row=len(data_effort_str)+6, column=1, value='Total')
    ws.cell(row=len(data_effort_str)+6, column=5, value=total_working)
    ws.cell(row=len(data_effort_str)+6, column=6, value=total_overtime)

    # write travel sheet
    ws2 = wb['travel']
    # print Title
    title = "Travel report for " + engineerName + ' of month ' + month
    ws2.cell(row=3, column=2, value=title)
    # print Header
    for i in range(0,len(header_travel)):
        ws2.cell(row=5, column=i+1, value=header_travel[i])

    # print Body
    for x in range(0,len(data_travel_str)):
        ws2.insert_rows(6+x)
        for y in range(1,len(header_travel)+1):
            ws2.cell(row=x+6, column=y, value=data_travel_str[x][y])

    # print Footer
    total_price = 0
    for item in data_travel_str:
        total_price += float(item[6])
    ws2.cell(row=len(data_travel_str)+6, column=1, value='Total')
    ws2.cell(row=len(data_travel_str)+6, column=5, value=total_price)

    filename = 'report' + engineerName + month + '.xlsx'
    print(filename)
    wb.save(filename)
    # return send_file('./excels/' + filename, attachment_filename=filename)

if __name__ == '__main__':
    # print(dictionary('package', 'PPE'))
    writeExcel('who', '08')